import json
import datetime
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.views import redirect_to_login
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, Max, Q, Sum, Avg
from django.db.models.functions import TruncDate, TruncHour
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_POST

from cafe_pos.models import (
    CafeTable,
    ChatAssistantSettings,
    ChatSession,
    Coupon,
    Customer,
    Floor,
    LoyaltySettings,
    Order,
    PaymentMethod,
    PaymentSettings,
    Product,
    ProductCategory,
    Profile,
    Promotion,
    ReceiptSettings,
    FeedbackQuestion,
    FeedbackResponse,
    OrderReview,
)
from tenants.models import AuditLog
from tenants.utils import log_action

from .forms import (
    CafeCustomizeForm,
    CategoryForm,
    ChatAssistantSettingsForm,
    CouponForm,
    EmployeeForm,
    FloorForm,
    CustomerLoyaltyForm,
    LoyaltySettingsForm,
    PaymentSettingsForm,
    ProductForm,
    PromotionForm,
    ReceiptSettingsForm,
    SetPasswordForm,
    TableForm,
)
from .mixins import cafe_admin_required


def _user_cafe(user):
    profile = getattr(user, "profile", None)
    if profile is not None and getattr(profile, "cafe_id", None):
        return profile.cafe
    return user.owned_cafes.first()


# ─────────────────────────────────────────────────────────────────────────────
# Host dispatcher
# ─────────────────────────────────────────────────────────────────────────────
def home(request):
    """Route the root URL based on the host: admin / public landing / cafe panel."""
    if getattr(request, "is_admin_host", False):
        return redirect("/admin/")
    if getattr(request, "cafe", None) is None:
        return render(request, "landing.html")
    # Cashiers go straight to the POS terminal; admins/superusers get the admin panel.
    if request.user.is_authenticated and not request.user.is_superuser:
        profile = getattr(request.user, "profile", None)
        if profile is not None and profile.cafe_id == request.cafe.id and profile.role == "cashier":
            return redirect("pos:terminal")
    return cafe_dashboard(request)


from cafe_pos.models import OrderLineItem

@cafe_admin_required(require_admin=False)
def cafe_dashboard(request):
    cafe = request.cafe
    today = timezone.localtime().date()
    start_of_month = today.replace(day=1)
    
    # 1. Today's Revenue and Orders
    today_orders = Order.objects.filter(cafe=cafe, created_at__date=today, status=Order.OrderStatus.PAID)
    today_revenue = today_orders.aggregate(total=Sum('total'))['total'] or 0
    today_count = today_orders.count()
    
    # 2. Monthly Revenue
    month_orders = Order.objects.filter(cafe=cafe, created_at__date__gte=start_of_month, status=Order.OrderStatus.PAID)
    monthly_revenue = month_orders.aggregate(total=Sum('total'))['total'] or 0
    
    # 3. Last 7 Days Revenue Trend
    seven_days_ago = today - datetime.timedelta(days=6)
    recent_orders = Order.objects.filter(
        cafe=cafe, 
        created_at__date__gte=seven_days_ago,
        status=Order.OrderStatus.PAID
    ).annotate(date=TruncDate('created_at')).values('date').annotate(revenue=Sum('total')).order_by('date')
    
    # Fill in missing days
    revenue_trend = []
    trend_dict = {str(item['date']): float(item['revenue']) for item in recent_orders}
    for i in range(7):
        d = seven_days_ago + datetime.timedelta(days=i)
        revenue_trend.append({
            "date": d.strftime("%b %d"),
            "revenue": trend_dict.get(str(d), 0.0)
        })

    # 4. Top 5 Products by Quantity Sold (Paid orders only)
    top_products = Product.objects.filter(cafe=cafe, order_lines__order__status=Order.OrderStatus.PAID) \
        .annotate(sold=Sum('order_lines__quantity')) \
        .exclude(sold=None) \
        .order_by('-sold')[:5]
        
    top_products_data = [{"name": p.name, "sold": p.sold} for p in top_products]

    # 5. Order Status Breakdown
    all_recent = Order.objects.filter(cafe=cafe, created_at__date__gte=start_of_month)
    status_counts = all_recent.values('status').annotate(count=Count('id'))
    status_dict = {item['status']: item['count'] for item in status_counts}
    order_stats = {
        "paid": status_dict.get(Order.OrderStatus.PAID, 0),
        "cancelled": status_dict.get(Order.OrderStatus.CANCELLED, 0),
        "draft": status_dict.get(Order.OrderStatus.DRAFT, 0),
    }

    context = {
        "today_revenue": today_revenue,
        "today_count": today_count,
        "monthly_revenue": monthly_revenue,
        "revenue_trend_json": json.dumps(revenue_trend),
        "top_products_json": json.dumps(top_products_data),
        "order_stats_json": json.dumps(order_stats),
        
        "stat_team": Profile.objects.filter(cafe=cafe, is_archived=False).count(),
        "stat_tables": CafeTable.objects.filter(cafe=cafe).count(),
        "recent_audit": AuditLog.objects.filter(cafe=cafe)[:8],
    }
    return render(request, "dashboard/index.html", context)


# ─────────────────────────────────────────────────────────────────────────────
# Floors & tables  (the Table View screen)
# ─────────────────────────────────────────────────────────────────────────────
@cafe_admin_required
def floors(request):
    cafe = request.cafe
    floor_list = list(Floor.objects.filter(cafe=cafe))
    active_id = request.GET.get("floor")
    active_floor = None
    if active_id:
        active_floor = next((f for f in floor_list if str(f.id) == active_id), None)
    if active_floor is None and floor_list:
        active_floor = floor_list[0]

    tables = []
    occupied_table_ids = set()
    if active_floor is not None:
        tables = list(CafeTable.objects.filter(floor=active_floor))
        from cafe_pos.models import Order as _Order
        occupied_table_ids = set(
            _Order.objects.filter(
                cafe=cafe, table__floor=active_floor,
                status__in=["draft", "sent_to_kitchen", "ready"]
            ).values_list("table_id", flat=True)
        )

    edit_table = None
    edit_table_id = request.GET.get("edit_table")
    if edit_table_id:
        edit_table = get_object_or_404(CafeTable, pk=edit_table_id, cafe=cafe)

    context = {
        "floors": floor_list,
        "active_floor": active_floor,
        "tables": tables,
        "occupied_table_ids": occupied_table_ids,
        "edit_floor": request.GET.get("edit_floor") == "1",
        "edit_table": edit_table,
        "floor_form": FloorForm(instance=active_floor),
        "table_form": TableForm(cafe=cafe, instance=edit_table),
    }
    return render(request, "dashboard/floors.html", context)


@cafe_admin_required
@require_POST
def floor_create(request):
    form = FloorForm(request.POST)
    if form.is_valid():
        floor = form.save(commit=False)
        floor.cafe = request.cafe
        floor.canvas_mode = request.POST.get("canvas_mode") == "on"
        floor.save()
        log_action("create", cafe=request.cafe, request=request, target=floor,
                   message=f"Added floor '{floor.name}'.")
        messages.success(request, f"Floor '{floor.name}' added.")
        return redirect(f"{_floors_url(request)}?floor={floor.id}")
    messages.error(request, "Could not add floor. Please check the name.")
    return redirect("dashboard:floors")


@cafe_admin_required
@require_POST
def floor_update(request, pk):
    floor = get_object_or_404(Floor, pk=pk, cafe=request.cafe)
    form = FloorForm(request.POST, instance=floor)
    if form.is_valid():
        floor = form.save(commit=False)
        # canvas_mode cannot be toggled here anymore, as per requirements
        floor.save()
        log_action("update", cafe=request.cafe, request=request, target=floor,
                   message=f"Renamed floor to '{floor.name}'.")
        messages.success(request, "Floor updated.")
    return redirect(f"{_floors_url(request)}?floor={floor.id}")

@cafe_admin_required
@require_POST
def save_floor_plan(request, pk):
    floor = get_object_or_404(Floor, pk=pk, cafe=request.cafe)
    try:
        data = json.loads(request.body)
        tables_data = data.get("tables", [])
        for t_data in tables_data:
            table_id = t_data.get("id")
            if not table_id: continue
            
            CafeTable.objects.filter(id=table_id, floor=floor).update(
                pos_x=t_data.get("pos_x", 0),
                pos_y=t_data.get("pos_y", 0),
                width=t_data.get("width", 80),
                height=t_data.get("height", 80),
                shape=t_data.get("shape", "rect")
            )
        return JsonResponse({"ok": True})
    except Exception as e:
        return JsonResponse({"ok": False, "message": str(e)})


@cafe_admin_required
@require_POST
def floor_delete(request, pk):
    floor = get_object_or_404(Floor, pk=pk, cafe=request.cafe)
    name = floor.name
    floor.delete()
    log_action("delete", cafe=request.cafe, request=request, target=None,
               message=f"Deleted floor '{name}' and its tables.")
    messages.success(request, f"Floor '{name}' deleted.")
    return redirect("dashboard:floors")


@cafe_admin_required
@require_POST
def table_create(request):
    form = TableForm(request.POST, cafe=request.cafe)
    if form.is_valid():
        table = form.save(commit=False)
        # Guard: the chosen floor must belong to this cafe.
        if table.floor.cafe_id != request.cafe.id:
            raise PermissionDenied("Invalid floor.")
        table.cafe = request.cafe
        table.sort_order = _next_table_sort_order(request.cafe, table.floor)
        table.save()
        log_action("create", cafe=request.cafe, request=request, target=table,
                   message=f"Added table '{table.table_number}' on {table.floor.name}.")
        messages.success(request, f"Table {table.table_number} added.")
        return redirect(f"{_floors_url(request)}?floor={table.floor_id}")
    messages.error(request, "Could not add table: " + "; ".join(
        f"{k}: {', '.join(v)}" for k, v in form.errors.items()))
    floor_id = request.POST.get("floor", "")
    return redirect(f"{_floors_url(request)}?floor={floor_id}")


@cafe_admin_required
@require_POST
def table_create_ajax(request):
    try:
        data = json.loads(request.body)
        floor = get_object_or_404(Floor, pk=data.get("floor_id"), cafe=request.cafe)
        table = CafeTable.objects.create(
            cafe=request.cafe,
            floor=floor,
            table_number=data.get("table_number"),
            seats=int(data.get("seats", 4)),
            pos_x=float(data.get("pos_x", 0)),
            pos_y=float(data.get("pos_y", 0)),
            width=float(data.get("width", 80)),
            height=float(data.get("height", 80)),
            shape="rect",
            sort_order=_next_table_sort_order(request.cafe, floor)
        )
        return JsonResponse({
            "ok": True,
            "table": {
                "id": table.id,
                "table_number": table.table_number,
                "seats": table.seats,
                "pos_x": table.pos_x,
                "pos_y": table.pos_y,
                "width": table.width,
                "height": table.height,
                "shape": table.shape
            }
        })
    except Exception as e:
        return JsonResponse({"ok": False, "message": str(e)})


@cafe_admin_required
@require_POST
def table_update(request, pk):
    table = get_object_or_404(CafeTable, pk=pk, cafe=request.cafe)
    previous_floor_id = table.floor_id
    form = TableForm(request.POST, instance=table, cafe=request.cafe)
    if form.is_valid():
        table = form.save(commit=False)
        table.save()
        if table.floor_id != previous_floor_id:
            table.sort_order = _next_table_sort_order(request.cafe, table.floor, exclude_table_id=table.pk)
            table.save(update_fields=["sort_order"])
        log_action("update", cafe=request.cafe, request=request, target=table,
                   message=f"Updated table '{table.table_number}'.")
        messages.success(request, "Table updated.")
    else:
        messages.error(request, "Could not update table.")
    return redirect(f"{_floors_url(request)}?floor={table.floor_id}")


@cafe_admin_required
@require_POST
def table_move(request):
    cafe = request.cafe
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "message": "Invalid payload."}, status=400)

    try:
        table_id = int(payload.get("table_id"))
        target_floor_id = int(payload.get("target_floor_id"))
    except (TypeError, ValueError):
        return JsonResponse({"ok": False, "message": "table_id and target_floor_id must be valid IDs."}, status=400)

    swap_table_id = payload.get("swap_table_id") or payload.get("before_table_id")
    try:
        swap_table_id = int(swap_table_id) if swap_table_id else None
    except (TypeError, ValueError):
        return JsonResponse({"ok": False, "message": "swap_table_id must be a valid ID."}, status=400)

    if not table_id or not target_floor_id:
        return JsonResponse({"ok": False, "message": "table_id and target_floor_id are required."}, status=400)

    table = get_object_or_404(CafeTable, pk=table_id, cafe=cafe)
    target_floor = get_object_or_404(Floor, pk=target_floor_id, cafe=cafe)
    swap_table = None
    if swap_table_id:
        swap_table = CafeTable.objects.filter(pk=swap_table_id, cafe=cafe, floor=target_floor).first()
        if swap_table is None:
            return JsonResponse({"ok": False, "message": "Target table was not found."}, status=404)

    with transaction.atomic():
        if swap_table is not None:
            source_floor = table.floor
            source_sort_order = table.sort_order

            table.floor = swap_table.floor
            table.sort_order = swap_table.sort_order
            swap_table.floor = source_floor
            swap_table.sort_order = source_sort_order

            if table.floor_id == swap_table.floor_id:
                table.save(update_fields=["sort_order"])
                swap_table.save(update_fields=["sort_order"])
                _resequence_floor_tables(cafe, table.floor_id)
            else:
                table.save(update_fields=["floor", "sort_order"])
                swap_table.save(update_fields=["floor", "sort_order"])
                _resequence_floor_tables(cafe, source_floor.id)
                _resequence_floor_tables(cafe, target_floor.id)

            from django.urls import reverse
            return JsonResponse({
                "ok": True,
                "floor_id": target_floor.id,
                "redirect_url": f"{reverse('dashboard:floors')}?floor={target_floor.id}",
            })

        source_floor_id = table.floor_id
        source_ids = list(
            CafeTable.objects.filter(cafe=cafe, floor_id=source_floor_id)
            .exclude(pk=table.pk)
            .order_by("sort_order", "id")
            .values_list("id", flat=True)
        )

        target_ids = []
        if target_floor_id == source_floor_id:
            target_ids = source_ids.copy()
        else:
            target_ids = list(
                CafeTable.objects.filter(cafe=cafe, floor_id=target_floor_id)
                .order_by("sort_order", "id")
                .values_list("id", flat=True)
            )

        target_ids.append(table.id)

        source_id_set = set(source_ids)
        target_id_set = set(target_ids)

        if source_floor_id == target_floor_id:
            moved_ids = target_ids
            tables_by_id = {
                item.id: item
                for item in CafeTable.objects.filter(cafe=cafe, id__in=moved_ids)
            }
            for sort_order, table_id_value in enumerate(moved_ids):
                row = tables_by_id[table_id_value]
                row.sort_order = sort_order
                row.save(update_fields=["sort_order"])
        else:
            source_tables_by_id = {
                item.id: item
                for item in CafeTable.objects.filter(cafe=cafe, id__in=source_id_set)
            }
            for sort_order, table_id_value in enumerate(source_ids):
                row = source_tables_by_id[table_id_value]
                row.sort_order = sort_order
                row.save(update_fields=["sort_order"])

            target_tables_by_id = {
                item.id: item
                for item in CafeTable.objects.filter(cafe=cafe, id__in=target_id_set)
            }
            for sort_order, table_id_value in enumerate(target_ids):
                row = target_tables_by_id[table_id_value]
                row.sort_order = sort_order
                if row.id == table.id:
                    row.floor = target_floor
                    row.save(update_fields=["floor", "sort_order"])
                else:
                    row.save(update_fields=["sort_order"])

    from django.urls import reverse
    return JsonResponse({
        "ok": True,
        "floor_id": target_floor.id,
        "redirect_url": f"{reverse('dashboard:floors')}?floor={target_floor.id}",
    })


@cafe_admin_required
@require_POST
def table_delete(request, pk):
    table = get_object_or_404(CafeTable, pk=pk, cafe=request.cafe)
    floor_id = table.floor_id
    number = table.table_number
    table.delete()
    _resequence_floor_tables(request.cafe, floor_id)
    log_action("delete", cafe=request.cafe, request=request, target=None,
               message=f"Deleted table '{number}'.")
    messages.success(request, f"Table {number} deleted.")
    return redirect(f"{_floors_url(request)}?floor={floor_id}")


def _floors_url(request):
    from django.urls import reverse
    return reverse("dashboard:floors")


def _next_table_sort_order(cafe, floor, exclude_table_id=None):
    queryset = CafeTable.objects.filter(cafe=cafe, floor=floor)
    if exclude_table_id is not None:
        queryset = queryset.exclude(pk=exclude_table_id)
    return (queryset.aggregate(max_sort=Max("sort_order"))["max_sort"] or -1) + 1


def _resequence_floor_tables(cafe, floor_id):
    tables = list(CafeTable.objects.filter(cafe=cafe, floor_id=floor_id).order_by("sort_order", "id"))
    changed = False
    for index, table in enumerate(tables):
        if table.sort_order != index:
            table.sort_order = index
            changed = True
    if changed:
        CafeTable.objects.bulk_update(tables, ["sort_order"])
@cafe_admin_required
@require_POST
def table_move(request):
    try:
        data = json.loads(request.body)
        table_id = data.get("table_id")
        target_floor_id = data.get("target_floor_id")
        swap_table_id = data.get("swap_table_id")

        table = get_object_or_404(CafeTable, pk=table_id, cafe=request.cafe)
        target_floor = get_object_or_404(Floor, pk=target_floor_id, cafe=request.cafe)

        # Move to target floor
        if table.floor_id != target_floor.id:
            table.floor = target_floor
            table.save(update_fields=["floor"])
            log_action("update", cafe=request.cafe, request=request, target=table,
                       message=f"Moved table '{table.table_number}' to {target_floor.name}.")

        if swap_table_id:
            swap_table = get_object_or_404(CafeTable, pk=swap_table_id, cafe=request.cafe)
            table.sort_order, swap_table.sort_order = swap_table.sort_order, table.sort_order
            table.save(update_fields=["sort_order"])
            swap_table.save(update_fields=["sort_order"])

        return JsonResponse({"ok": True, "redirect_url": f"{_floors_url(request)}?floor={target_floor.id}"})
    except Exception as e:
        return JsonResponse({"ok": False, "message": str(e)}, status=400)


# ─────────────────────────────────────────────────────────────────────────────
# Products & categories
# ─────────────────────────────────────────────────────────────────────────────
@cafe_admin_required
def products(request):
    cafe = request.cafe
    product_list = Product.objects.filter(cafe=cafe).select_related("category")
    return render(request, "dashboard/products.html", {
        "products": product_list,
        "has_categories": ProductCategory.objects.filter(cafe=cafe).exists(),
    })


@cafe_admin_required
def product_create(request):
    cafe = request.cafe
    if request.method == "POST":
        form = ProductForm(request.POST, request.FILES, cafe=cafe)
        if form.is_valid():
            product = form.save(commit=False)
            product.cafe = cafe
            product.save()
            log_action("create", cafe=cafe, request=request, target=product,
                       message=f"Added product '{product.name}'.")
            messages.success(request, f"Product '{product.name}' added.")
            return redirect("dashboard:products")
    else:
        form = ProductForm(cafe=cafe)
    return render(request, "dashboard/product_form.html", {"form": form, "mode": "Add"})


@cafe_admin_required
def product_update(request, pk):
    cafe = request.cafe
    product = get_object_or_404(Product, pk=pk, cafe=cafe)
    if request.method == "POST":
        form = ProductForm(request.POST, request.FILES, instance=product, cafe=cafe)
        if form.is_valid():
            form.save()
            log_action("update", cafe=cafe, request=request, target=product,
                       message=f"Updated product '{product.name}'.")
            messages.success(request, "Product updated.")
            return redirect("dashboard:products")
    else:
        form = ProductForm(instance=product, cafe=cafe)
    return render(request, "dashboard/product_form.html", {"form": form, "mode": "Edit", "product": product})


@cafe_admin_required
@require_POST
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk, cafe=request.cafe)
    name = product.name
    product.delete()
    log_action("delete", cafe=request.cafe, request=request, target=None,
               message=f"Deleted product '{name}'.")
    messages.success(request, f"Product '{name}' deleted.")
    return redirect("dashboard:products")


@cafe_admin_required
@require_POST
def product_bulk_action(request):
    action = request.POST.get("action")
    product_ids = request.POST.getlist("product_ids")
    if not action or not product_ids:
        messages.error(request, "No action or products selected.")
        return redirect("dashboard:products")
    
    products = Product.objects.filter(cafe=request.cafe, id__in=product_ids)
    count = products.count()
    if action == "delete":
        products.delete()
        log_action("delete", cafe=request.cafe, request=request, target=None,
                   message=f"Deleted {count} products in bulk.")
        messages.success(request, f"{count} products deleted.")
    elif action == "archive":
        products.update(is_active=False)
        log_action("update", cafe=request.cafe, request=request, target=None,
                   message=f"Archived {count} products in bulk.")
        messages.success(request, f"{count} products archived.")
    else:
        messages.error(request, "Invalid action.")
    return redirect("dashboard:products")


@cafe_admin_required
def categories(request):
    cafe = request.cafe
    edit_id = request.GET.get("edit")
    instance = None
    if edit_id:
        instance = get_object_or_404(ProductCategory, pk=edit_id, cafe=cafe)
    form = CategoryForm(instance=instance)
    return render(request, "dashboard/categories.html", {
        "categories": ProductCategory.objects.filter(cafe=cafe),
        "form": form,
        "edit_instance": instance,
    })


@cafe_admin_required
@require_POST
def category_create(request):
    cafe = request.cafe
    form = CategoryForm(request.POST)
    if form.is_valid():
        category = form.save(commit=False)
        category.cafe = cafe
        category.save()
        log_action("create", cafe=cafe, request=request, target=category,
                   message=f"Added category '{category.name}'.")
        messages.success(request, f"Category '{category.name}' added.")
    else:
        messages.error(request, "Could not add category (name may already exist).")
    return redirect("dashboard:categories")


@cafe_admin_required
@require_POST
def category_update(request, pk):
    category = get_object_or_404(ProductCategory, pk=pk, cafe=request.cafe)
    form = CategoryForm(request.POST, instance=category)
    if form.is_valid():
        form.save()
        log_action("update", cafe=request.cafe, request=request, target=category,
                   message=f"Updated category '{category.name}'.")
        messages.success(request, "Category updated.")
    return redirect("dashboard:categories")


@cafe_admin_required
@require_POST
def category_delete(request, pk):
    category = get_object_or_404(ProductCategory, pk=pk, cafe=request.cafe)
    name = category.name
    if category.products.exists():
        messages.error(request, f"Cannot delete '{name}' — it still has products.")
        return redirect("dashboard:categories")
    category.delete()
    log_action("delete", cafe=request.cafe, request=request, target=None,
               message=f"Deleted category '{name}'.")
    messages.success(request, f"Category '{name}' deleted.")
    return redirect("dashboard:categories")


# ─────────────────────────────────────────────────────────────────────────────
# Coupons & Promotions
# ─────────────────────────────────────────────────────────────────────────────
@cafe_admin_required
def coupons(request):
    cafe = request.cafe
    edit_coupon_id = request.GET.get("edit_coupon")
    edit_promo_id = request.GET.get("edit_promo")
    coupon_instance = get_object_or_404(Coupon, pk=edit_coupon_id, cafe=cafe) if edit_coupon_id else None
    promo_instance = get_object_or_404(Promotion, pk=edit_promo_id, cafe=cafe) if edit_promo_id else None

    coupon_list = Coupon.objects.filter(cafe=cafe).order_by("code")
    promo_list = Promotion.objects.filter(cafe=cafe).order_by("name")

    coupon_form = CouponForm(instance=coupon_instance, cafe=cafe)
    promo_form = PromotionForm(instance=promo_instance, cafe=cafe)

    return render(request, "dashboard/coupons.html", {
        "coupon_list": coupon_list,
        "promo_list": promo_list,
        "coupon_form": coupon_form,
        "promo_form": promo_form,
        "edit_coupon": coupon_instance,
        "edit_promo": promo_instance,
        "products": Product.objects.filter(cafe=cafe, is_active=True),
    })


@cafe_admin_required
@require_POST
def coupon_create(request):
    form = CouponForm(request.POST, cafe=request.cafe)
    if form.is_valid():
        coupon = form.save(commit=False)
        coupon.cafe = request.cafe
        coupon.save()
        log_action("create", cafe=request.cafe, request=request, target=coupon,
                   message=f"Created coupon '{coupon.code}'.")
        messages.success(request, f"Coupon '{coupon.code}' created.")
    else:
        messages.error(request, "Could not create coupon: " + "; ".join(
            f"{k}: {', '.join(v)}" for k, v in form.errors.items()))
    return redirect("dashboard:coupons")


@cafe_admin_required
@require_POST
def coupon_update(request, pk):
    coupon = get_object_or_404(Coupon, pk=pk, cafe=request.cafe)
    form = CouponForm(request.POST, instance=coupon, cafe=request.cafe)
    if form.is_valid():
        form.save()
        log_action("update", cafe=request.cafe, request=request, target=coupon,
                   message=f"Updated coupon '{coupon.code}'.")
        messages.success(request, "Coupon updated.")
    else:
        messages.error(request, "Could not update coupon.")
    return redirect("dashboard:coupons")


@cafe_admin_required
@require_POST
def coupon_delete(request, pk):
    coupon = get_object_or_404(Coupon, pk=pk, cafe=request.cafe)
    code = coupon.code
    coupon.delete()
    log_action("delete", cafe=request.cafe, request=request, target=None,
               message=f"Deleted coupon '{code}'.")
    messages.success(request, f"Coupon '{code}' deleted.")
    return redirect("dashboard:coupons")


@cafe_admin_required
@require_POST
def coupon_toggle(request, pk):
    coupon = get_object_or_404(Coupon, pk=pk, cafe=request.cafe)
    coupon.is_active = not coupon.is_active
    coupon.save(update_fields=["is_active"])
    log_action("update", cafe=request.cafe, request=request, target=coupon,
               message=f"Toggled coupon '{coupon.code}' {'active' if coupon.is_active else 'inactive'}.")
    return redirect("dashboard:coupons")


@cafe_admin_required
@require_POST
def promotion_create(request):
    form = PromotionForm(request.POST, cafe=request.cafe)
    if form.is_valid():
        promo = form.save(commit=False)
        promo.cafe = request.cafe
        promo.save()
        log_action("create", cafe=request.cafe, request=request, target=promo,
                   message=f"Created promotion '{promo.name}'.")
        messages.success(request, f"Promotion '{promo.name}' created.")
    else:
        messages.error(request, "Could not create promotion: " + "; ".join(
            f"{k}: {', '.join(v)}" for k, v in form.errors.items()))
    return redirect("dashboard:coupons")


@cafe_admin_required
@require_POST
def promotion_update(request, pk):
    promo = get_object_or_404(Promotion, pk=pk, cafe=request.cafe)
    form = PromotionForm(request.POST, instance=promo, cafe=request.cafe)
    if form.is_valid():
        form.save()
        log_action("update", cafe=request.cafe, request=request, target=promo,
                   message=f"Updated promotion '{promo.name}'.")
        messages.success(request, "Promotion updated.")
    else:
        messages.error(request, "Could not update promotion: " + "; ".join(
            f"{k}: {', '.join(v)}" for k, v in form.errors.items()))
    return redirect("dashboard:coupons")


@cafe_admin_required
@require_POST
def promotion_delete(request, pk):
    promo = get_object_or_404(Promotion, pk=pk, cafe=request.cafe)
    name = promo.name
    promo.delete()
    log_action("delete", cafe=request.cafe, request=request, target=None,
               message=f"Deleted promotion '{name}'.")
    messages.success(request, f"Promotion '{name}' deleted.")
    return redirect("dashboard:coupons")


@cafe_admin_required
@require_POST
def promotion_toggle(request, pk):
    promo = get_object_or_404(Promotion, pk=pk, cafe=request.cafe)
    promo.is_active = not promo.is_active
    promo.save(update_fields=["is_active"])
    log_action("update", cafe=request.cafe, request=request, target=promo,
               message=f"Toggled promotion '{promo.name}' {'active' if promo.is_active else 'inactive'}.")
    return redirect("dashboard:coupons")


# ─────────────────────────────────────────────────────────────────────────────
# Loyalty & customer ranking
# ─────────────────────────────────────────────────────────────────────────────
@cafe_admin_required
def loyalty(request):
    cafe = request.cafe
    settings_obj, _ = LoyaltySettings.objects.get_or_create(cafe=cafe)
    settings_form = LoyaltySettingsForm(instance=settings_obj)
    customers = (
        Customer.objects.filter(cafe=cafe)
        .annotate(paid_orders=Count("orders", filter=Q(orders__status=Order.OrderStatus.PAID)))
        .order_by("-paid_orders", "name")
    )
    rows = [
        {
            "customer": customer,
            "loyalty": customer.loyalty_snapshot(settings_obj=settings_obj, paid_orders=customer.paid_orders),
            "form": CustomerLoyaltyForm(instance=customer),
        }
        for customer in customers
    ]
    return render(request, "dashboard/loyalty.html", {
        "settings_form": settings_form,
        "settings": settings_obj,
        "customer_rows": rows,
    })


@cafe_admin_required
@require_POST
def loyalty_settings_update(request):
    settings_obj, _ = LoyaltySettings.objects.get_or_create(cafe=request.cafe)
    form = LoyaltySettingsForm(request.POST, instance=settings_obj)
    if form.is_valid():
        form.save()
        log_action("update", cafe=request.cafe, request=request, target=settings_obj,
                   message="Updated loyalty settings.")
        messages.success(request, "Loyalty settings saved.")
    else:
        messages.error(request, "Could not save loyalty settings: " + "; ".join(
            f"{k}: {', '.join(v)}" for k, v in form.errors.items()))
    return redirect("dashboard:loyalty")


@cafe_admin_required
@require_POST
def customer_loyalty_update(request, pk):
    customer = get_object_or_404(Customer, pk=pk, cafe=request.cafe)
    form = CustomerLoyaltyForm(request.POST, instance=customer)
    if form.is_valid():
        form.save()
        log_action("update", cafe=request.cafe, request=request, target=customer,
                   message=f"Updated loyalty status for '{customer.name}'.")
        messages.success(request, f"Customer '{customer.name}' updated.")
    else:
        messages.error(request, "Could not update customer.")
    return redirect("dashboard:loyalty")


# ─────────────────────────────────────────────────────────────────────────────
# Team / employees
# ─────────────────────────────────────────────────────────────────────────────
@cafe_admin_required
def users(request):
    cafe = request.cafe
    form = EmployeeForm()
    if request.method == "POST":
        form = EmployeeForm(request.POST)
        if form.is_valid():
            User = get_user_model()
            user = User.objects.create_user(
                username=form.cleaned_data["username"],
                email=form.cleaned_data["email"],
                password=form.cleaned_data["password"],
            )
            Profile.objects.update_or_create(
                user=user,
                defaults={"cafe": cafe, "role": form.cleaned_data["role"], "is_archived": False},
            )
            log_action("create", cafe=cafe, request=request, target=user,
                       message=f"Added team member '{user.username}' ({form.cleaned_data['role']}).")
            messages.success(request, f"Team member '{user.username}' added.")
            return redirect("dashboard:users")
    members = Profile.objects.filter(cafe=cafe).select_related("user").order_by("is_archived", "user__username")
    return render(request, "dashboard/users.html", {"members": members, "form": form})


@cafe_admin_required
@require_POST
def user_archive(request, pk):
    profile = get_object_or_404(Profile, pk=pk, cafe=request.cafe)
    if profile.user_id == request.user.id:
        messages.error(request, "You cannot archive your own account.")
        return redirect("dashboard:users")
    profile.is_archived = not profile.is_archived
    profile.user.is_active = not profile.is_archived
    profile.user.save(update_fields=["is_active"])
    profile.save(update_fields=["is_archived"])
    log_action("update", cafe=request.cafe, request=request, target=profile.user,
               message=f"{'Archived' if profile.is_archived else 'Restored'} '{profile.user.username}'.")
    messages.success(request, f"{'Archived' if profile.is_archived else 'Restored'} {profile.user.username}.")
    return redirect("dashboard:users")


@cafe_admin_required
@require_POST
def user_password(request, pk):
    profile = get_object_or_404(Profile, pk=pk, cafe=request.cafe)
    form = SetPasswordForm(request.POST)
    if form.is_valid():
        profile.user.set_password(form.cleaned_data["password"])
        profile.user.save(update_fields=["password"])
        log_action("update", cafe=request.cafe, request=request, target=profile.user,
                   message=f"Reset password for '{profile.user.username}'.")
        messages.success(request, f"Password updated for {profile.user.username}.")
    else:
        messages.error(request, "Password too weak (min 8 characters).")
    return redirect("dashboard:users")


@cafe_admin_required
@require_POST
def user_delete(request, pk):
    profile = get_object_or_404(Profile, pk=pk, cafe=request.cafe)
    if profile.user_id == request.user.id:
        messages.error(request, "You cannot delete your own account.")
        return redirect("dashboard:users")
    username = profile.user.username
    profile.user.delete()
    log_action("delete", cafe=request.cafe, request=request, target=None,
               message=f"Deleted team member '{username}'.")
    messages.success(request, f"'{username}' has been deleted.")
    return redirect("dashboard:users")


# ─────────────────────────────────────────────────────────────────────────────
# Customize cafe
# ─────────────────────────────────────────────────────────────────────────────
@cafe_admin_required
def customize(request):
    cafe = request.cafe
    if request.method == "POST":
        form = CafeCustomizeForm(request.POST, request.FILES, instance=cafe)
        if form.is_valid():
            form.save()
            log_action("update", cafe=cafe, request=request, target=cafe,
                       message="Updated cafe branding/details.")
            messages.success(request, "Cafe details saved.")
            return redirect("dashboard:customize")
    else:
        form = CafeCustomizeForm(instance=cafe)
    return render(request, "dashboard/customize.html", {"form": form})


# ─────────────────────────────────────────────────────────────────────────────
# Audit log
# ─────────────────────────────────────────────────────────────────────────────
@cafe_admin_required
def audit_log(request):
    logs = AuditLog.objects.filter(cafe=request.cafe).select_related("actor")
    paginator = Paginator(logs, 30)
    page = paginator.get_page(request.GET.get("page"))
    return render(request, "dashboard/audit_log.html", {"page_obj": page})


# ─────────────────────────────────────────────────────────────────────────────
# Payment settings (Razorpay / Stripe / UPI config)
# ─────────────────────────────────────────────────────────────────────────────
@cafe_admin_required
def payment_settings(request):
    from pos.services import ensure_payment_methods

    cafe = request.cafe
    ensure_payment_methods(cafe)
    settings_obj, _ = PaymentSettings.objects.get_or_create(cafe=cafe)
    if request.method == "POST":
        form = PaymentSettingsForm(request.POST, instance=settings_obj)
        if form.is_valid():
            form.save()
            for t in ("cash", "card", "upi"):
                pm = PaymentMethod.objects.filter(cafe=cafe, type=t).first()
                if pm:
                    pm.is_enabled = bool(request.POST.get(f"enable_{t}"))
                    if t == "upi" and form.cleaned_data.get("upi_id"):
                        pm.upi_id = form.cleaned_data["upi_id"]
                    pm.save()
            log_action("update", cafe=cafe, request=request, target=settings_obj,
                       message="Updated payment settings.")
            messages.success(request, "Payment settings saved.")
            return redirect("dashboard:payment-methods")
    else:
        form = PaymentSettingsForm(instance=settings_obj)
    methods = {m.type: m for m in PaymentMethod.objects.filter(cafe=cafe)}
    return render(request, "dashboard/payment_settings.html", {"form": form, "methods": methods})


# ─────────────────────────────────────────────────────────────────────────────
# Receipts (default + custom HTML with data pills + per-cafe SMTP)
# ─────────────────────────────────────────────────────────────────────────────
@cafe_admin_required
def receipt_settings(request):
    from cafe_pos.receipts import DATA_PILLS

    cafe = request.cafe
    rs, _ = ReceiptSettings.objects.get_or_create(cafe=cafe)
    if request.method == "POST":
        form = ReceiptSettingsForm(request.POST, instance=rs)
        if form.is_valid():
            form.save()
            log_action("update", cafe=cafe, request=request, target=rs,
                       message="Updated receipt settings.")
            messages.success(request, "Receipt settings saved.")
            return redirect("dashboard:receipts")
    else:
        form = ReceiptSettingsForm(instance=rs)
    return render(request, "dashboard/receipt_settings.html",
                  {"form": form, "pills": DATA_PILLS, "settings": rs})


@cafe_admin_required
@require_POST
def receipt_preview(request):
    """Render the receipt against the latest order, using the POSTed (unsaved) HTML."""
    from cafe_pos.receipts import order_context
    from django.template import Context, Template
    from django.template.loader import render_to_string
    from django.utils.safestring import mark_safe

    order = Order.objects.filter(cafe=request.cafe).order_by("-created_at").first()
    if order is None:
        return HttpResponse("<p style='padding:24px;font-family:sans-serif;color:#888'>"
                            "No orders yet — take one in the POS to preview a receipt.</p>")
    html_template = (request.POST.get("template_html") or "").strip()
    use_default = request.POST.get("use_default") == "on" or not html_template
    if use_default:
        return HttpResponse(render_to_string("receipts/default_receipt.html", order_context(order)))
    ctx = order_context(order)
    ctx["items_table"] = mark_safe(ctx["items_table"])
    ctx["logo"] = mark_safe(ctx["logo"])
    try:
        html = Template(html_template).render(Context(ctx))
    except Exception as exc:  # show template errors inline in the preview
        html = f"<p style='color:#c0392b;padding:24px;font-family:sans-serif'>Template error: {exc}</p>"
    return HttpResponse(html)


@cafe_admin_required
@require_POST
def receipt_test(request):
    from cafe_pos.receipts import email_receipt

    cafe = request.cafe
    order = Order.objects.filter(cafe=cafe).order_by("-created_at").first()
    to = (request.POST.get("email") or request.user.email or "").strip()
    if order is None:
        messages.error(request, "No orders yet to use as a sample receipt.")
    elif not to:
        messages.error(request, "Add an email address to send the test to.")
    elif email_receipt(order, to):
        messages.success(request, f"Test receipt sent to {to}.")
    else:
        messages.error(request, "Could not send the test — check the SMTP settings.")
    return redirect("dashboard:receipts")


RECEIPT_THEMES = {
    "minimal": {"name": "Clean Minimal", "desc": "White, line-based, lightweight. Great for takeaway.", "file": "minimal.html"},
    "dark_luxury": {"name": "Dark Luxury", "desc": "Charcoal & gold, serif typography. Fine dining.", "file": "dark_luxury.html"},
    "cafe_warm": {"name": "Cafe Warm", "desc": "Warm brown tones, logo-centric. Classic café feel.", "file": "cafe_warm.html"},
    "modern_card": {"name": "Modern Card", "desc": "Gradient header, card layout, vibrant purple.", "file": "modern_card.html"},
}


@cafe_admin_required
@require_POST
def receipt_apply_theme(request):
    """Return the HTML content of a receipt theme so the editor can load it."""
    import json
    from pathlib import Path

    data = json.loads(request.body) if request.content_type == "application/json" else {}
    slug = data.get("theme", "")
    if slug not in RECEIPT_THEMES:
        return JsonResponse({"error": "Unknown theme."}, status=400)
    theme_path = Path(settings.BASE_DIR) / "templates" / "receipts" / "themes" / RECEIPT_THEMES[slug]["file"]
    if not theme_path.is_file():
        return JsonResponse({"error": "Theme file not found."}, status=404)
    html = theme_path.read_text(encoding="utf-8")
    return JsonResponse({"html": html, "name": RECEIPT_THEMES[slug]["name"]})


@cafe_admin_required
def receipt_themes_list(request):
    """Return the list of available themes as JSON."""
    themes = [
        {"slug": slug, "name": info["name"], "desc": info["desc"]}
        for slug, info in RECEIPT_THEMES.items()
    ]
    return JsonResponse({"themes": themes})


# ─────────────────────────────────────────────────────────────────────────────
# Reports
# ─────────────────────────────────────────────────────────────────────────────
def _reports_queryset(cafe, date_from, date_to, user_id=None, session_id=None, product_id=None):
    """Paid orders in the period, with optional filters."""
    qs = Order.objects.filter(
        cafe=cafe,
        status=Order.OrderStatus.PAID,
        paid_at__date__gte=date_from,
        paid_at__date__lte=date_to,
    )
    if user_id:
        qs = qs.filter(employee_id=user_id)
    if session_id:
        qs = qs.filter(session_id=session_id)
    if product_id:
        qs = qs.filter(line_items__product_id=product_id)
    return qs.distinct()


@cafe_admin_required
def reports(request):
    from django.contrib.auth import get_user_model
    User = get_user_model()
    cafe = request.cafe

    today = timezone.localdate()
    raw_from = request.GET.get("date_from")
    raw_to = request.GET.get("date_to")
    date_from = parse_date(raw_from) if raw_from else (today - datetime.timedelta(days=6))
    date_to = parse_date(raw_to) if raw_to else today
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    user_id = request.GET.get("user_id") or None
    session_id = request.GET.get("session_id") or None
    product_id = request.GET.get("product_id") or None

    qs = _reports_queryset(cafe, date_from, date_to, user_id, session_id, product_id)

    # ── KPIs ─────────────────────────────────────────────────────────────────
    agg = qs.aggregate(orders=Count("id"), revenue=Sum("total"), avg_order=Avg("total"))
    total_orders = agg["orders"] or 0
    total_revenue = float(agg["revenue"] or 0)
    avg_order = float(agg["avg_order"] or 0)

    # Previous period for change %
    period_days = (date_to - date_from).days + 1
    prev_to = date_from - datetime.timedelta(days=1)
    prev_from = prev_to - datetime.timedelta(days=period_days - 1)
    qs_prev = _reports_queryset(cafe, prev_from, prev_to, user_id, session_id, product_id)
    agg_prev = qs_prev.aggregate(orders=Count("id"), revenue=Sum("total"))
    prev_orders = agg_prev["orders"] or 0
    prev_revenue = float(agg_prev["revenue"] or 0)

    def pct_change(now, prev):
        if prev == 0:
            return None
        return round((now - prev) / prev * 100, 1)

    # ── Sales chart ──────────────────────────────────────────────────────────
    if period_days <= 2:
        sales_qs = qs.annotate(bucket=TruncHour("paid_at")).values("bucket").annotate(
            rev=Sum("total"), cnt=Count("id")
        ).order_by("bucket")
        chart_labels = [r["bucket"].strftime("%I %p") for r in sales_qs]
    else:
        sales_qs = qs.annotate(bucket=TruncDate("paid_at")).values("bucket").annotate(
            rev=Sum("total"), cnt=Count("id")
        ).order_by("bucket")
        chart_labels = [r["bucket"].strftime("%d %b") for r in sales_qs]
    chart_revenue = [float(r["rev"]) for r in sales_qs]

    # ── Category pie ─────────────────────────────────────────────────────────
    from cafe_pos.models import OrderLineItem, ProductCategory
    cat_data = (
        OrderLineItem.objects.filter(order__in=qs)
        .values("product__category__name")
        .annotate(cnt=Sum("quantity"), rev=Sum("line_total"))
        .order_by("-rev")[:8]
    )
    cat_labels = [r["product__category__name"] or "Uncategorized" for r in cat_data]
    cat_values = [float(r["rev"]) for r in cat_data]

    # ── Top orders ───────────────────────────────────────────────────────────
    top_orders = (
        qs.select_related("table", "customer", "employee", "session")
        .order_by("-total")[:10]
    )

    # ── Top products ─────────────────────────────────────────────────────────
    top_products = (
        OrderLineItem.objects.filter(order__in=qs)
        .values("product__name")
        .annotate(qty=Sum("quantity"), rev=Sum("line_total"))
        .order_by("-rev")[:10]
    )

    # ── Top categories ───────────────────────────────────────────────────────
    top_categories = (
        OrderLineItem.objects.filter(order__in=qs)
        .values("product__category__name")
        .annotate(rev=Sum("line_total"))
        .order_by("-rev")[:10]
    )

    team = cafe.profiles.filter(is_archived=False).select_related("user")
    sessions = cafe.pos_sessions.order_by("-opened_at")[:20]
    products = Product.objects.filter(cafe=cafe, is_active=True)

    return render(request, "dashboard/reports.html", {
        "date_from": date_from,
        "date_to": date_to,
        "user_id": user_id or "",
        "session_id": session_id or "",
        "product_id": product_id or "",
        "total_orders": total_orders,
        "total_revenue": total_revenue,
        "avg_order": avg_order,
        "orders_change": pct_change(total_orders, prev_orders),
        "revenue_change": pct_change(total_revenue, prev_revenue),
        "chart_labels": json.dumps(chart_labels),
        "chart_revenue": json.dumps(chart_revenue),
        "cat_labels": json.dumps(cat_labels),
        "cat_values": json.dumps(cat_values),
        "top_orders": top_orders,
        "top_products": top_products,
        "top_categories": top_categories,
        "team": team,
        "sessions": sessions,
        "products": products,
        "period_days": period_days,
    })


@cafe_admin_required
def reports_export(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    cafe = request.cafe
    today = timezone.localdate()
    raw_from = request.GET.get("date_from")
    raw_to = request.GET.get("date_to")
    date_from = parse_date(raw_from) if raw_from else (today - datetime.timedelta(days=6))
    date_to = parse_date(raw_to) if raw_to else today
    user_id = request.GET.get("user_id") or None
    session_id = request.GET.get("session_id") or None
    product_id = request.GET.get("product_id") or None

    qs = _reports_queryset(cafe, date_from, date_to, user_id, session_id, product_id)
    from cafe_pos.models import OrderLineItem

    wb = openpyxl.Workbook()
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="C8903E")
    hdr_align = Alignment(horizontal="center")

    def add_sheet(name, headers, rows):
        ws = wb.create_sheet(name)
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = hdr_align
        for row_idx, row in enumerate(rows, 2):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
        return ws

    orders_rows = [
        (o.order_number, o.session_id, o.paid_at.strftime("%d/%m/%Y %H:%M") if o.paid_at else "",
         o.customer.name if o.customer_id else "", o.employee.username,
         float(o.total))
        for o in qs.select_related("customer", "employee").order_by("-paid_at")
    ]
    add_sheet("Orders", ["Order #", "Session", "Date", "Customer", "Employee", "Total (₹)"], orders_rows)

    top_products = (
        OrderLineItem.objects.filter(order__in=qs)
        .values("product__name").annotate(qty=Sum("quantity"), rev=Sum("line_total")).order_by("-rev")
    )
    add_sheet("Top Products", ["Product", "Qty", "Revenue (₹)"],
              [(r["product__name"], r["qty"], float(r["rev"])) for r in top_products])

    top_cats = (
        OrderLineItem.objects.filter(order__in=qs)
        .values("product__category__name").annotate(rev=Sum("line_total")).order_by("-rev")
    )
    add_sheet("Top Categories", ["Category", "Revenue (₹)"],
              [(r["product__category__name"] or "Uncategorized", float(r["rev"])) for r in top_cats])

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"report_{date_from}_{date_to}.xlsx"
    response = HttpResponse(buf.read(),
                            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    return response



@cafe_admin_required
def feedback_settings(request):
    cafe = request.cafe
    receipt_settings, _ = ReceiptSettings.objects.get_or_create(cafe=cafe)
    
    if request.method == "POST":
        receipt_settings.feedback_email_html = request.POST.get("feedback_email_html", "")
        receipt_settings.save(update_fields=["feedback_email_html"])
        
        # Handle dynamic questions
        q_ids = request.POST.getlist("question_id")
        q_texts = request.POST.getlist("question_text")
        q_types = request.POST.getlist("question_type")
        
        # Delete existing ones not in the submitted list
        existing_ids = [int(i) for i in q_ids if i.isdigit()]
        FeedbackQuestion.objects.filter(cafe=cafe).exclude(id__in=existing_ids).delete()
        
        for idx, text in enumerate(q_texts):
            text = text.strip()
            if not text:
                continue
            qid = q_ids[idx] if idx < len(q_ids) else ""
            qtype = q_types[idx] if idx < len(q_types) else "rating"
            
            if qid.isdigit():
                q = FeedbackQuestion.objects.get(id=int(qid), cafe=cafe)
                q.question_text = text
                q.type = qtype
                q.sort_order = idx
                q.save()
            else:
                FeedbackQuestion.objects.create(
                    cafe=cafe, question_text=text, type=qtype, sort_order=idx
                )
                
        messages.success(request, "Feedback settings updated successfully.")
        return redirect("dashboard:feedback-settings")
        
    questions = FeedbackQuestion.objects.filter(cafe=cafe).order_by("sort_order")
    
    return render(request, "dashboard/feedback_settings.html", {
        "receipt_settings": receipt_settings,
        "questions": questions,
    })

@cafe_admin_required
def feedback_report(request):
    cafe = request.cafe
    reviews = OrderReview.objects.filter(cafe=cafe).select_related("order", "customer", "cashier").prefetch_related("responses__question", "kitchen_staff").order_by("-created_at")
    
    # Analytics
    total_reviews = reviews.count()
    avg_rating = reviews.aggregate(Avg('rating'))['rating__avg'] or 0.0
    
    # Rating Distribution (1-5 stars)
    distribution = list(reviews.values('rating').annotate(count=Count('id')).order_by('rating'))
    dist_dict = {item['rating']: item['count'] for item in distribution}
    rating_distribution = [
        {"rating": i, "count": dist_dict.get(i, 0)} for i in range(1, 6)
    ]
    
    # Per-question averages
    question_stats = []
    from cafe_pos.models import FeedbackQuestion, FeedbackResponse
    for q in FeedbackQuestion.objects.filter(cafe=cafe, type='rating'):
        avg = FeedbackResponse.objects.filter(review__cafe=cafe, question=q).aggregate(Avg('rating_value'))['rating_value__avg'] or 0
        question_stats.append({
            "question": q.question_text,
            "avg": avg
        })
    
    paginator = Paginator(reviews, 50)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    
    return render(request, "dashboard/feedback_report.html", {
        "page_obj": page_obj,
        "total_reviews": total_reviews,
        "avg_rating": avg_rating,
        "rating_distribution_json": json.dumps(rating_distribution),
        "question_stats_json": json.dumps(question_stats),
    })


@cafe_admin_required
@require_POST
def feedback_preview(request):
    from cafe_pos.receipts import order_context
    from django.template import Context, Template
    from django.template.loader import render_to_string
    from django.utils.safestring import mark_safe
    
    order = Order.objects.filter(cafe=request.cafe).order_by("-created_at").first()
    if order is None:
        return HttpResponse("<p style='padding:24px;font-family:sans-serif;color:#888'>No orders yet — take one in the POS to preview.</p>")
        
    html_template = (request.POST.get("template_html") or request.POST.get("feedback_email_html") or "").strip()
    ctx = order_context(order)
    ctx["items_table"] = mark_safe(ctx.get("items_table", ""))
    ctx["logo"] = mark_safe(ctx.get("logo", ""))
    
    if not html_template:
        html = f'''
        <div style="font-family: sans-serif; max-width: 500px; margin: 0 auto; text-align: center;">
            <h2>Thank you for visiting {order.cafe.name}!</h2>
            <p>We hope you enjoyed your order (<b>#{order.order_number}</b>).</p>
            <p>Please take a moment to leave us a review. Your feedback helps us improve!</p>
            <a href="{ctx.get('review_url', '#')}" style="display: inline-block; padding: 12px 24px; background: #c8903e; color: #fff; text-decoration: none; border-radius: 8px; font-weight: bold; margin-top: 10px;">Leave a Review</a>
        </div>
        '''
        return HttpResponse(html)
        
    try:
        html = Template(html_template).render(Context(ctx))
    except Exception as exc:
        html = f"<p style='color:#c0392b;padding:24px;font-family:sans-serif'>Template error: {exc}</p>"
    return HttpResponse(html)

# ─────────────────────────────────────────────────────────────────────────────
# Placeholders for deferred POS modules
# ─────────────────────────────────────────────────────────────────────────────
@cafe_admin_required(require_admin=False)
def kds_display(request):
    return redirect("pos:kds")


@cafe_admin_required(require_admin=False)
def coming_soon(request):
    label = {
        "pos-session": "POS Terminal",
        "kds": "Kitchen Display",
        "reports": "Reports",
        "bookings": "Bookings",
        "payment-methods": "Payment Methods",
        "coupons": "Coupons & Promotions",
    }.get(request.resolver_match.url_name, "This module")
    return render(request, "dashboard/coming_soon.html", {"label": label})


# ─────────────────────────────────────────────────────────────────────────────
# AI Chat Assistant
# ─────────────────────────────────────────────────────────────────────────────

@cafe_admin_required
def assistant_settings(request):
    import json
    cafe = request.cafe
    obj, _ = ChatAssistantSettings.objects.get_or_create(cafe=cafe)

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "scrape":
            from cafe_pos.chatbot import scrape_menu_data
            from django.utils import timezone as tz
            data = scrape_menu_data(cafe)
            obj.product_data_json = json.dumps(data, ensure_ascii=False, indent=2)
            obj.last_scraped_at = tz.now()
            obj.save(update_fields=["product_data_json", "last_scraped_at"])
            messages.success(request, "Menu data refreshed successfully.")
            return redirect("dashboard:assistant-settings")

        form = ChatAssistantSettingsForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            log_action("update", cafe=cafe, request=request, message="Updated AI chat assistant settings.")
            messages.success(request, "Chat assistant settings saved.")
            return redirect("dashboard:assistant-settings")
    else:
        form = ChatAssistantSettingsForm(instance=obj)

    menu_summary = None
    if obj.product_data_json:
        try:
            parsed = json.loads(obj.product_data_json)
            cats = parsed.get("menu", [])
            total_items = sum(len(c.get("items", [])) for c in cats)
            menu_summary = {"categories": len(cats), "items": total_items}
        except Exception:
            pass

    return render(request, "dashboard/assistant_settings.html", {
        "form": form,
        "obj": obj,
        "menu_summary": menu_summary,
    })


@cafe_admin_required
def assistant_sessions(request):
    cafe = request.cafe
    qs = ChatSession.objects.filter(cafe=cafe).prefetch_related("messages").select_related("order")
    page = Paginator(qs, 25).get_page(request.GET.get("page"))
    return render(request, "dashboard/assistant_sessions.html", {"page": page})


@cafe_admin_required
def assistant_session_detail(request, pk):
    session = get_object_or_404(ChatSession, pk=pk, cafe=request.cafe)
    msgs = session.messages.order_by("created_at")
    return render(request, "dashboard/assistant_session_detail.html", {
        "session": session,
        "msgs": msgs,
    })


@cafe_admin_required
@require_POST
def assistant_session_delete(request, pk):
    session = get_object_or_404(ChatSession, pk=pk, cafe=request.cafe)
    session.delete()
    messages.success(request, "Conversation deleted.")
    return redirect("dashboard:assistant-sessions")
